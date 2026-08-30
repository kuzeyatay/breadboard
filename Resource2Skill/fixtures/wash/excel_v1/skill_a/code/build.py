"""Quarterly P&L workbook scaffold (deterministic, no external state)."""
from openpyxl import Workbook


def render(workbook_path: str = "quarterly_pnl.xlsx") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Q1", "Q2", "Q3", "Q4"])
    ws.append(["Revenue", 120000, 135000, 148000, 162000])
    ws.append(["COGS", 70000, 78000, 85000, 92000])
    ws.append(["Gross Profit", "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3"])
    wb.create_sheet("Details")
    wb.create_sheet("Charts")
    wb.save(workbook_path)
    return workbook_path
