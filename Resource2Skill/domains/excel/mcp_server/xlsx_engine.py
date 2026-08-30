"""
domains/excel/mcp_server/xlsx_engine.py

In-memory openpyxl execution engine for the Excel domain MCP server.

Responsibilities
----------------
- Maintain a registry of live ``openpyxl.Workbook`` objects keyed by id
- Run user-supplied openpyxl code against those workbooks inside an isolated
  temp CWD so stray ``output.xlsx`` files do not pollute the project
- Load design tokens (themes / format presets / chart templates) from JSON
- Load and dispatch tier-tagged skill modules (components / sheet shells /
  workbook archetypes) by importing their ``.py`` files from the skill library
- Verify a saved .xlsx file by re-opening it and reporting structural metadata
"""
from __future__ import annotations

import importlib.util
import io
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
import uuid
from pathlib import Path
from typing import Any, Callable

# openpyxl is the only hard dep. Imported lazily so the module can still be
# inspected in environments without openpyxl installed.
_openpyxl = None
_chart_mod = None
_styles_mod = None
_utils_mod = None
_formatting_mod = None
_table_mod = None


def _ox():
    global _openpyxl, _chart_mod, _styles_mod, _utils_mod, _formatting_mod, _table_mod
    if _openpyxl is None:
        import openpyxl
        from openpyxl import chart as chart_mod
        from openpyxl import styles as styles_mod
        from openpyxl import utils as utils_mod
        from openpyxl import formatting as formatting_mod
        from openpyxl.worksheet import table as table_mod
        _openpyxl = openpyxl
        _chart_mod = chart_mod
        _styles_mod = styles_mod
        _utils_mod = utils_mod
        _formatting_mod = formatting_mod
        _table_mod = table_mod
    return _openpyxl


# ---------------------------------------------------------------------------
# Workbook registry
# ---------------------------------------------------------------------------

_workbooks: dict[str, Any] = {}
_workbook_meta: dict[str, dict] = {}


def new_workbook_id() -> str:
    return f"wb_{uuid.uuid4().hex[:8]}"


def create_workbook(name: str | None = None, theme: str | None = None) -> str:
    ox = _ox()
    wb = ox.Workbook()
    # Drop the default empty sheet — caller will add named sheets explicitly.
    # We still keep one placeholder to satisfy openpyxl's "must have a sheet"
    # invariant; it gets removed on first add_sheet.
    wb_id = new_workbook_id()
    _workbooks[wb_id] = wb
    _workbook_meta[wb_id] = {
        "name": name or "workbook",
        "theme": theme,
        "created_sheets": [],
    }
    return wb_id


def get_workbook(workbook_id: str):
    if workbook_id not in _workbooks:
        raise KeyError(f"Unknown workbook_id: {workbook_id}")
    return _workbooks[workbook_id]


def get_meta(workbook_id: str) -> dict:
    return _workbook_meta.setdefault(workbook_id, {})


def drop_workbook(workbook_id: str) -> None:
    _workbooks.pop(workbook_id, None)
    _workbook_meta.pop(workbook_id, None)


def list_workbooks() -> list[dict]:
    out = []
    for wid, wb in _workbooks.items():
        meta = _workbook_meta.get(wid, {})
        out.append({
            "workbook_id": wid,
            "name": meta.get("name"),
            "theme": meta.get("theme"),
            "sheets": [s.title for s in wb.worksheets],
        })
    return out


# ---------------------------------------------------------------------------
# Path safety
# ---------------------------------------------------------------------------

_SHELL_METACHARS = re.compile(r"[;|&$`<>]")


def safe_demo_path(filepath: str, demo_root: Path) -> Path:
    """Resolve ``filepath`` and confirm it lives under ``demo_root``.

    Rejects shell metacharacters and ``..`` traversal via realpath comparison.
    """
    if _SHELL_METACHARS.search(filepath):
        raise ValueError(f"filepath contains shell metacharacters: {filepath!r}")
    candidate = Path(filepath)
    if not candidate.is_absolute():
        candidate = demo_root / candidate.name
    resolved = Path(os.path.realpath(candidate))
    demo_resolved = Path(os.path.realpath(demo_root))
    try:
        resolved.relative_to(demo_resolved)
    except ValueError:
        raise ValueError(
            f"filepath {filepath!r} resolves outside demo dir {demo_resolved}"
        )
    return resolved


# ---------------------------------------------------------------------------
# Save / verify
# ---------------------------------------------------------------------------

def save_workbook(workbook_id: str, filepath: str, demo_root: Path) -> dict:
    wb = get_workbook(workbook_id)
    target = safe_demo_path(filepath, demo_root)
    target.parent.mkdir(parents=True, exist_ok=True)
    saved_cwd = os.getcwd()
    tmp = tempfile.mkdtemp(prefix="xlsx_save_")
    try:
        os.chdir(tmp)
        wb.save(str(target))
    finally:
        os.chdir(saved_cwd)
        shutil.rmtree(tmp, ignore_errors=True)
    return {
        "workbook_id": workbook_id,
        "filepath": str(target),
        "size_bytes": target.stat().st_size,
        "sheets": [s.title for s in wb.worksheets],
    }


def verify_workbook(filepath: str, demo_root: Path) -> dict:
    """Re-open the .xlsx and return structural metadata."""
    target = safe_demo_path(filepath, demo_root)
    if not target.exists():
        return {"ok": False, "error": f"file not found: {target}"}
    ox = _ox()
    wb = ox.load_workbook(str(target), data_only=False)
    sheets = []
    warnings: list[str] = []
    for ws in wb.worksheets:
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        non_empty = 0
        if rows and cols:
            for row in ws.iter_rows(min_row=1, max_row=min(rows, 200),
                                    max_col=cols, values_only=True):
                if any(c is not None and c != "" for c in row):
                    non_empty += 1
        chart_count = len(getattr(ws, "_charts", []) or [])
        table_count = len(getattr(ws, "tables", {}) or {})
        if rows == 1 and cols == 1 and ws.cell(1, 1).value in (None, ""):
            warnings.append(f"sheet '{ws.title}' is empty")
        sheets.append({
            "name": ws.title,
            "rows": rows,
            "cols": cols,
            "non_empty_rows": non_empty,
            "charts": chart_count,
            "tables": table_count,
        })
    return {
        "ok": True,
        "filepath": str(target),
        "size_bytes": target.stat().st_size,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Token loading (themes / format_presets / chart_templates / formulas)
# ---------------------------------------------------------------------------

def _read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _legacy_skills_dir() -> Path:
    """Fallback dir for token JSONs (themes/format_presets/chart_templates/formulas).
    The wiki backend doesn't ship these design tokens — fall back to the
    legacy seed library so the agent's color/format calls keep working.
    """
    return Path(__file__).resolve().parents[3] / "skills_library" / "excel"


def list_tokens(skills_dir: Path, kind: str) -> list[str]:
    base = skills_dir / kind
    names: set[str] = set()
    if base.exists():
        names.update(p.stem for p in base.glob("*.json"))
    legacy = _legacy_skills_dir() / kind
    if legacy.exists() and legacy != base:
        names.update(p.stem for p in legacy.glob("*.json"))
    return sorted(names)


def get_token(skills_dir: Path, kind: str, name: str) -> dict:
    p = skills_dir / kind / f"{name}.json"
    if p.exists():
        return _read_json(p)
    legacy = _legacy_skills_dir() / kind / f"{name}.json"
    if legacy.exists():
        return _read_json(legacy)
    raise FileNotFoundError(f"{kind}/{name}.json not found in {skills_dir} or legacy")


# ---------------------------------------------------------------------------
# Skill module loading (components / sheet_shells / archetypes / formulas)
# ---------------------------------------------------------------------------

_module_cache: dict[str, Any] = {}


def _load_skill_module(py_path: Path):
    key = str(py_path.resolve())
    if key in _module_cache:
        return _module_cache[key]
    spec = importlib.util.spec_from_file_location(
        f"excel_skill_{py_path.stem}", py_path
    )
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot load {py_path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    _module_cache[key] = mod
    return mod


def _resolve_skill_path(skills_dir: Path, tier_dir: str, skill_id: str) -> Path:
    """Find the skill's executable .py module, supporting BOTH layouts.

    Legacy layout: <skills_dir>/<tier_dir>/<skill_id>.py
    Wiki layout:   <skills_dir>/<skill_id>/code/skill.py  (flat, tier in meta.json)
    """
    # Legacy layout (tier-grouped subdirs)
    base = skills_dir / tier_dir
    direct = base / f"{skill_id}.py"
    if direct.exists():
        return direct
    # Wiki layout (flat skill_id directories with code/skill.py)
    wiki_path = skills_dir / skill_id / "code" / "skill.py"
    if wiki_path.exists():
        return wiki_path
    # Distinguish "skill exists in library but is reference-only" from
    # "skill_id unknown" so the caller can give a useful error.
    index_path = skills_dir / "index.json"
    skill_in_index = False
    if index_path.exists():
        try:
            idx = json.loads(index_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            idx = None
        if idx:
            for s in idx.get("skills", []):
                if s.get("skill_id") == skill_id:
                    skill_in_index = True
                    break
    if skill_in_index:
        raise FileNotFoundError(
            f"skill {skill_id!r} is a distilled (reference-only) "
            f"skill with no executable .py module. Read its technique "
            f"with get_skill_code({skill_id!r}) and adapt the code "
            f"inside execute_xlsx_code instead."
        )
    raise FileNotFoundError(f"skill {skill_id} not found in {base}")


def apply_component(skills_dir: Path, workbook_id: str, sheet_name: str,
                    anchor: str, component_id: str, kwargs: dict | None = None):
    wb = get_workbook(workbook_id)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    py = _resolve_skill_path(skills_dir, "components", component_id)
    mod = _load_skill_module(py)
    fn: Callable | None = getattr(mod, "render", None)
    if fn is None:
        raise AttributeError(f"component {component_id} missing render(ws, anchor, **kw)")
    fn(ws, anchor, **(kwargs or {}))
    return {"ok": True, "component": component_id, "sheet": sheet_name, "anchor": anchor}


def render_sheet_shell(skills_dir: Path, workbook_id: str, shell_id: str,
                       sheet_name: str, kwargs: dict | None = None):
    wb = get_workbook(workbook_id)
    py = _resolve_skill_path(skills_dir, "sheet_shells_seed", shell_id)
    mod = _load_skill_module(py)
    fn: Callable | None = getattr(mod, "render_sheet", None)
    if fn is None:
        raise AttributeError(f"shell {shell_id} missing render_sheet(wb, sheet_name, **kw)")
    fn(wb, sheet_name, **(kwargs or {}))
    meta = get_meta(workbook_id)
    meta.setdefault("created_sheets", []).append({"name": sheet_name, "shell": shell_id})
    return {"ok": True, "shell": shell_id, "sheet": sheet_name}


def render_archetype(skills_dir: Path, workbook_id: str, archetype_id: str,
                     kwargs: dict | None = None):
    wb = get_workbook(workbook_id)
    py = _resolve_skill_path(skills_dir, "workbook_archetypes", archetype_id)
    mod = _load_skill_module(py)
    fn: Callable | None = getattr(mod, "render_workbook", None)
    if fn is None:
        raise AttributeError(
            f"archetype {archetype_id} missing render_workbook(wb, **kw)"
        )
    fn(wb, **(kwargs or {}))
    meta = get_meta(workbook_id)
    meta["archetype"] = archetype_id
    return {
        "ok": True,
        "archetype": archetype_id,
        "sheets": [s.title for s in wb.worksheets],
    }


# ---------------------------------------------------------------------------
# Free code execution (the escape hatch)
# ---------------------------------------------------------------------------

def execute_code(workbook_id: str, code: str, skills_dir: Path) -> dict:
    """Execute openpyxl code against the live workbook in an isolated CWD.

    The local namespace pre-populates ``wb``, ``ox``, common style classes,
    chart classes, and a ``token(kind, name)`` helper.
    """
    wb = get_workbook(workbook_id)
    ox = _ox()
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle, Color,
        GradientFill, Protection,
    )
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, AreaChart, Reference,
        BarChart3D, LineChart3D,
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import (
        CellIsRule, FormulaRule, ColorScaleRule, IconSetRule, DataBarRule,
    )
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.formatting import Rule
    from copy import copy as _copy

    ns = {
        "wb": wb,
        "ox": ox,
        "Font": Font, "PatternFill": PatternFill, "Border": Border,
        "Side": Side, "Alignment": Alignment, "NamedStyle": NamedStyle,
        "Color": Color, "GradientFill": GradientFill, "Protection": Protection,
        "BarChart": BarChart, "LineChart": LineChart, "PieChart": PieChart,
        "ScatterChart": ScatterChart, "AreaChart": AreaChart,
        "Reference": Reference, "BarChart3D": BarChart3D,
        "LineChart3D": LineChart3D, "DataLabelList": DataLabelList,
        "CellIsRule": CellIsRule, "FormulaRule": FormulaRule,
        "ColorScaleRule": ColorScaleRule, "IconSetRule": IconSetRule,
        "DataBarRule": DataBarRule,
        "get_column_letter": get_column_letter,
        "column_index_from_string": column_index_from_string,
        "Table": Table, "TableStyleInfo": TableStyleInfo,
        "Rule": Rule,
        "copy_style": _copy,
        "token": lambda kind, name: get_token(skills_dir, kind, name),
    }

    saved_cwd = os.getcwd()
    tmp = tempfile.mkdtemp(prefix="xlsx_exec_")
    stdout_buf = io.StringIO()
    saved_stdout = sys.stdout
    try:
        os.chdir(tmp)
        sys.stdout = stdout_buf
        exec(compile(code, "<execute_xlsx_code>", "exec"), ns, ns)
        return {
            "ok": True,
            "stdout": stdout_buf.getvalue(),
            "sheet_names": [s.title for s in wb.worksheets],
        }
    except Exception as e:
        return {
            "ok": False,
            "error": f"{type(e).__name__}: {e}",
            "traceback": traceback.format_exc(),
            "stdout": stdout_buf.getvalue(),
        }
    finally:
        sys.stdout = saved_stdout
        os.chdir(saved_cwd)
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------------------
# Convenience: workbook info
# ---------------------------------------------------------------------------

def workbook_info(workbook_id: str) -> dict:
    wb = get_workbook(workbook_id)
    meta = get_meta(workbook_id)
    sheets = []
    for ws in wb.worksheets:
        sheets.append({
            "name": ws.title,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "charts": len(getattr(ws, "_charts", []) or []),
            "tables": list((getattr(ws, "tables", {}) or {}).keys()),
        })
    return {
        "workbook_id": workbook_id,
        "name": meta.get("name"),
        "theme": meta.get("theme"),
        "archetype": meta.get("archetype"),
        "sheets": sheets,
    }
