"""
domains/excel/mcp_server/server.py
Excel MCP Server — exposes tier-based tools for openpyxl-native workbook
construction.

Tool tiers
----------
T1 Tokens     get_palette_preset / get_format_preset / get_chart_template
              / get_formula_snippet / list_tokens
T2 Snippets   get_formula_snippet (also used as snippet store)
T3 Components apply_component
T4 Shells     add_sheet_from_shell
T5 Archetype  init_from_archetype
Workbook      create_workbook / save_workbook / get_workbook_info /
              verify_workbook
Free          execute_xlsx_code
Library       list_skills / get_skill_info / get_skill_code

Launch:
    python domains/excel/mcp_server/server.py --skills-dir skills_library/excel
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

_SERVER_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _SERVER_DIR.parents[2]
sys.path.insert(0, str(_PROJECT_ROOT))
sys.path.insert(0, str(_SERVER_DIR))

from mcp.server.fastmcp import FastMCP

import xlsx_engine as engine
from core.skill_grounding import artifact_manifest_path, make_grounding_entries, write_manifest

log = logging.getLogger("excel-mcp")

mcp = FastMCP("excel-agent")

_skills_dir: Path | None = None
_demo_root: Path | None = None
_index_cache: dict | None = None


def _skills() -> Path:
    if _skills_dir is None:
        raise RuntimeError("--skills-dir not configured")
    return _skills_dir


def _demo() -> Path:
    if _demo_root is None:
        return _PROJECT_ROOT / "demo" / "excel"
    return _demo_root


def _load_index() -> dict:
    global _index_cache
    if _index_cache is not None:
        return _index_cache
    p = _skills() / "index.json"
    if not p.exists():
        _index_cache = {"domain": "excel", "total": 0, "skills": []}
        return _index_cache
    with p.open("r", encoding="utf-8") as f:
        _index_cache = json.load(f)
    return _index_cache


def _invalidate_index():
    global _index_cache
    _index_cache = None


def _ok(payload: dict) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _err(msg: str) -> str:
    return f"Error: {msg}"


def _append_workbook_groundings(
    workbook_id: str,
    *,
    tool_name: str,
    from_skill_ids,
    target_node: str = "",
    adaptation_notes: str = "",
    fallback_target: str = "",
    extra: dict | None = None,
) -> list[dict]:
    try:
        meta = engine.get_meta(workbook_id)
    except Exception:
        return []
    entries = make_grounding_entries(
        domain="excel",
        tool_name=tool_name,
        from_skill_ids=from_skill_ids,
        target_node=target_node,
        adaptation_notes=adaptation_notes,
        fallback_target=fallback_target,
        extra=extra,
    )
    if entries:
        meta.setdefault("skill_groundings", []).extend(entries)
    return entries


def _write_workbook_manifest(workbook_id: str, filepath: str) -> str:
    meta = engine.get_meta(workbook_id)
    manifest_path = artifact_manifest_path(filepath)
    write_manifest(
        manifest_path,
        {
            "domain": "excel",
            "workbook_id": workbook_id,
            "workbook_name": meta.get("name"),
            "theme": meta.get("theme"),
            "groundings": meta.get("skill_groundings", []),
        },
        domain="excel",
    )
    return str(manifest_path)


# ---------------------------------------------------------------------------
# Registry cache invalidation + backend-flip recovery
# ---------------------------------------------------------------------------

_startup_backend: str | None = None
_wiki_tools_registered: bool = False


def _rebuild_runtime() -> dict:
    """Re-resolve _skills_dir from the current library_backend, refresh the
    legacy index cache, and toggle the universal wiki tool surface to match.

    This is the operator-facing recovery path for AC-13: a backend flip in
    ``domain.yaml`` followed by a ``reload_registry`` call must move the
    served registry without restarting the process.
    """
    global _skills_dir, _wiki_tools_registered
    info: dict = {"reloaded": True}

    try:
        from core import get_active_library_backend, get_library_dir
        backend = get_active_library_backend("excel")
        _skills_dir = get_library_dir("excel").resolve()
    except Exception as exc:  # noqa: BLE001
        info["error"] = f"backend resolve failed: {type(exc).__name__}: {exc}"
        return info

    info["backend"] = backend
    info["skills_dir"] = str(_skills_dir)
    _invalidate_index()

    # Match the runtime tool surface to the new backend.
    try:
        from domains.excel.wiki_adapter import ExcelWikiAdapter
        from core.skill_wiki.mcp_tools import register_wiki_tools, _OWNED_NAMES
        from core.skill_wiki.legacy_stale import mark_runtime_backend
    except ImportError as exc:
        info["adapter_error"] = f"adapter import failed: {exc}"
        return info

    rebuild_succeeded = False
    if backend == "wiki":
        try:
            adapter = ExcelWikiAdapter()
            adapter.reload()
            register_wiki_tools(mcp, adapter)
            _wiki_tools_registered = True
            info["adapter_reloaded"] = True
            info["tool_surface"] = "wiki"
            rebuild_succeeded = True
        except Exception as exc:  # noqa: BLE001
            info["adapter_reloaded"] = False
            info["adapter_error"] = f"{type(exc).__name__}: {exc}"
    else:
        # Backend is legacy. If wiki tools were previously registered,
        # drop them so the legacy surface is canonical again.
        if _wiki_tools_registered:
            manager = getattr(mcp, "_tool_manager", None)
            tools = getattr(manager, "_tools", {}) if manager else {}
            for name in _OWNED_NAMES:
                tools.pop(name, None)
            _wiki_tools_registered = False
        info["tool_surface"] = "legacy"
        rebuild_succeeded = True

    # Only re-key the stale guard if the rebuild succeeded. A failed wiki
    # registration must leave guarded legacy tools returning stale_registry
    # so operators are not silently unstuck under a mismatched backend.
    if rebuild_succeeded:
        mark_runtime_backend(mcp, backend)
    else:
        info["stale_guard_left_armed"] = True
    return info


@mcp.tool()
def reload_registry() -> dict:
    """Recover after a ``library_backend`` flip or external registry update.

    Re-reads the configured backend, recomputes the served skills directory,
    invalidates legacy caches, and rebuilds the MCP tool surface to match.
    This tool is intentionally exempt from the legacy stale guard so it
    remains callable after a flip — it is the documented remediation path.
    """
    return _rebuild_runtime()


# ---------------------------------------------------------------------------
# Workbook lifecycle
# ---------------------------------------------------------------------------

@mcp.tool()
def create_workbook(name: str = "workbook", theme: str = "") -> str:
    """Create an in-memory workbook. Returns workbook_id (use it for all
    subsequent calls). The workbook stays in memory until save_workbook.

    Args:
        name: human label (e.g. "quarterly_report")
        theme: optional palette name (must exist in themes/), used by shells
               and components when not explicitly overridden
    """
    try:
        wid = engine.create_workbook(name=name, theme=theme or None)
        engine.get_meta(wid).setdefault("skill_groundings", [])
    except Exception as e:
        return _err(f"create_workbook failed: {e}")
    return _ok({"workbook_id": wid, "name": name, "theme": theme or None})


@mcp.tool()
def save_workbook(workbook_id: str, filepath: str) -> str:
    """Save the workbook to disk under demo/excel/. Path must resolve inside
    demo/excel/ — symlink/`..` traversal is rejected."""
    try:
        info = engine.save_workbook(workbook_id, filepath, _demo())
        info["skill_trace_manifest"] = _write_workbook_manifest(
            workbook_id, info["filepath"]
        )
    except Exception as e:
        return _err(f"save_workbook failed: {e}")
    return _ok(info)


@mcp.tool()
def get_workbook_info(workbook_id: str) -> str:
    """Return sheet names, row/col counts, chart/table counts for a live
    workbook (in-memory, before save)."""
    try:
        return _ok(engine.workbook_info(workbook_id))
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def verify_workbook(filepath: str) -> str:
    """Re-open a saved .xlsx and return structural verification: sheet list,
    rows/cols, charts, tables, plus warnings (e.g. empty sheets)."""
    try:
        return _ok(engine.verify_workbook(filepath, _demo()))
    except Exception as e:
        return _err(f"verify_workbook failed: {e}")


@mcp.tool()
def workbook_reflect(workbook_id: str) -> str:
    """Rubric-aware self-audit before save. Returns per-sheet stats + warnings
    aimed at the same dimensions the workbook is scored on:

    - data_density: data rows per non-summary sheet (target ≥60)
    - structure_clarity: sheet count, named sheets, leftover default 'Sheet'
    - chart_quality: chart count, chart data ranges (empty range = blank box)
    - placeholder/lorem: counts cells equal to 'foo','bar','TBD','Lorem',etc.

    Use this AFTER you have populated all sheets but BEFORE save_workbook.
    If the report flags issues (e.g. <60 data rows on a sheet that should
    have source data), go fix them before saving."""
    try:
        from openpyxl.utils import get_column_letter as _gcl
        wb = engine._workbooks.get(workbook_id) if hasattr(engine, "_workbooks") else None
        if wb is None:
            try:
                wb_info = engine.workbook_info(workbook_id)
                wb = engine._workbooks.get(workbook_id)
            except Exception:
                pass
        if wb is None:
            return _err(f"workbook_id {workbook_id} not found")

        placeholder_tokens = {"foo","bar","baz","tbd","todo","lorem","ipsum","xxxx","example","placeholder","sample","item 1","item 2"}
        report = {"sheets": [], "warnings": [], "totals": {}}
        total_rows = 0
        total_charts = 0
        total_placeholders = 0
        empty_sheets = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            non_empty_cells = 0
            placeholder_hits = 0
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None or v == "":
                        continue
                    non_empty_cells += 1
                    if isinstance(v, str) and v.strip().lower() in placeholder_tokens:
                        placeholder_hits += 1
            data_rows = max(0, max_r - 1) if max_r > 1 else 0
            chart_count = len(getattr(ws, "_charts", []) or [])
            chart_ranges = []
            for ch in getattr(ws, "_charts", []) or []:
                try:
                    sers = [s.values.numRef.f for s in ch.series if s.values and getattr(s.values, "numRef", None)]
                    chart_ranges.extend(sers)
                except Exception:
                    pass

            sheet_rep = {
                "sheet": sheet_name,
                "rows": max_r,
                "cols": max_c,
                "data_rows": data_rows,
                "non_empty_cells": non_empty_cells,
                "charts": chart_count,
                "chart_ranges": chart_ranges,
                "placeholder_hits": placeholder_hits,
            }
            report["sheets"].append(sheet_rep)
            total_rows += data_rows
            total_charts += chart_count
            total_placeholders += placeholder_hits
            if non_empty_cells == 0:
                empty_sheets += 1
                report["warnings"].append(f"sheet '{sheet_name}' is empty")
            elif data_rows < 60 and sheet_name.lower() not in {"cover","summary","dashboard","kpi","kpis","overview"}:
                report["warnings"].append(f"sheet '{sheet_name}' has only {data_rows} data rows — target ≥60 for source-data sheets")
            if placeholder_hits > 0:
                report["warnings"].append(f"sheet '{sheet_name}' contains {placeholder_hits} placeholder strings (foo/bar/TBD/...)")

        report["totals"] = {
            "sheet_count": len(wb.sheetnames),
            "active_sheet": wb.active.title if wb.active is not None else "?",
            "total_data_rows": total_rows,
            "total_charts": total_charts,
            "total_placeholders": total_placeholders,
            "empty_sheets": empty_sheets,
        }
        if total_charts == 0:
            report["warnings"].append("no charts in workbook — add at least one BarChart/LineChart on summary sheet if brief mentions trend/dashboard/comparison")
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            report["warnings"].append("default 'Sheet' still present — remove via wb.remove(wb['Sheet'])")

        return _ok(report)
    except Exception as e:
        return _err(f"workbook_reflect failed: {e}")


# ---------------------------------------------------------------------------
# Free code execution (escape hatch)
# ---------------------------------------------------------------------------

@mcp.tool()
def execute_xlsx_code(
    workbook_id: str,
    code: str,
    from_skill_ids: str = "",
    target_node: str = "",
    adaptation_notes: str = "",
) -> str:
    """Run arbitrary openpyxl code against the live workbook.

    Pre-injected names: wb, ox, Font, PatternFill, Border, Side, Alignment,
    Color, GradientFill, BarChart, LineChart, PieChart, ScatterChart,
    AreaChart, BarChart3D, LineChart3D, Reference, DataLabelList,
    CellIsRule, FormulaRule, ColorScaleRule, IconSetRule, DataBarRule,
    Table, TableStyleInfo, Rule, get_column_letter, column_index_from_string,
    copy_style, token(kind, name).

    Use this for any openpyxl operation not covered by a dedicated tool.
    Code runs in an isolated temp CWD; stdout is captured and returned.
    from_skill_ids: optional JSON/comma list of inspected wiki skill ids whose
        mechanisms the generated code adapts.
    target_node: optional sheet/block/role target, or JSON/comma list for
        multiple grounded sections.
    adaptation_notes: optional note describing borrowed mechanisms.
    """
    try:
        out = engine.execute_code(workbook_id, code, _skills())
        if out.get("ok"):
            entries = _append_workbook_groundings(
                workbook_id,
                tool_name="execute_xlsx_code",
                from_skill_ids=from_skill_ids,
                target_node=target_node,
                adaptation_notes=adaptation_notes,
                fallback_target=target_node or "workbook_code",
                extra={"code_chars": len(code)},
            )
            if entries:
                out["grounded_sections"] = len(entries)
        return _ok(out)
    except Exception as e:
        return _err(str(e))


# ---------------------------------------------------------------------------
# Tier 1 / Tier 2: tokens
# ---------------------------------------------------------------------------

_TOKEN_KINDS = ("themes", "format_presets", "chart_templates", "formulas",
                "cell_techniques")


@mcp.tool()
def list_tokens(kind: str = "themes") -> str:
    """List token names available under skills_library/excel/<kind>/.

    kind ∈ {themes, format_presets, chart_templates, formulas, cell_techniques}
    """
    if kind not in _TOKEN_KINDS:
        return _err(f"unknown kind {kind!r}; expected one of {_TOKEN_KINDS}")
    return _ok({"kind": kind, "names": engine.list_tokens(_skills(), kind)})


@mcp.tool()
def get_palette_preset(name: str) -> str:
    """Return a theme palette JSON: header_bg, header_fg, zebra, accent,
    title_fg, body_fg, fonts, etc. Use for format_range bg_color/font_color."""
    try:
        return _ok(engine.get_token(_skills(), "themes", name))
    except FileNotFoundError as e:
        return _err(str(e))


@mcp.tool()
def get_format_preset(name: str) -> str:
    """Return a number-format preset JSON: number_format string + optional
    font/fill/alignment hints. Names: currency_usd, currency_signed,
    percent_signed, thousand_sep, kpi_traffic, date_compact, etc."""
    try:
        return _ok(engine.get_token(_skills(), "format_presets", name))
    except FileNotFoundError as e:
        return _err(str(e))


@mcp.tool()
def get_chart_template(name: str) -> str:
    """Return a chart template JSON: chart_type, style number, axis font
    sizes, color scheme, gridline settings."""
    try:
        return _ok(engine.get_token(_skills(), "chart_templates", name))
    except FileNotFoundError as e:
        return _err(str(e))


@mcp.tool()
def get_formula_snippet(name: str) -> str:
    """Return a formula template JSON: pattern (with {placeholders}),
    description, example. Use string-format inside execute_xlsx_code."""
    try:
        return _ok(engine.get_token(_skills(), "formulas", name))
    except FileNotFoundError as e:
        return _err(str(e))


# ---------------------------------------------------------------------------
# Tier 3 / Tier 4 / Tier 5: dispatchers
# ---------------------------------------------------------------------------

@mcp.tool()
def apply_component(workbook_id: str, sheet_name: str, anchor: str,
                    component_id: str, kwargs_json: str = "{}") -> str:
    """Apply a tier-3 component (small, sheet-local) at a cell anchor.

    Args:
        sheet_name: target sheet (created if missing)
        anchor:    top-left cell, e.g. "B2"
        component_id: file stem under skills_library/excel/components/
        kwargs_json: JSON dict passed to render(ws, anchor, **kwargs)
    """
    try:
        kwargs = json.loads(kwargs_json) if kwargs_json else {}
    except json.JSONDecodeError as e:
        return _err(f"kwargs_json invalid: {e}")
    try:
        out = engine.apply_component(_skills(), workbook_id, sheet_name,
                                     anchor, component_id, kwargs)
        _append_workbook_groundings(
            workbook_id,
            tool_name="apply_component",
            from_skill_ids=[component_id],
            target_node=f"{sheet_name}:{anchor}",
            adaptation_notes=f"component kwargs={kwargs_json[:240]}",
            extra={"sheet_name": sheet_name, "anchor": anchor},
        )
    except Exception as e:
        return _err(f"apply_component failed: {e}")
    return _ok(out)


@mcp.tool()
def add_sheet_from_shell(workbook_id: str, shell_id: str, sheet_name: str,
                        kwargs_json: str = "{}") -> str:
    """Add a tier-4 sheet shell (full sheet skeleton with theme + layout).

    The shell module's render_sheet(wb, sheet_name, **kwargs) builds the
    whole sheet. kwargs typically include data, theme, title.
    """
    try:
        kwargs = json.loads(kwargs_json) if kwargs_json else {}
    except json.JSONDecodeError as e:
        return _err(f"kwargs_json invalid: {e}")
    try:
        out = engine.render_sheet_shell(_skills(), workbook_id, shell_id,
                                        sheet_name, kwargs)
        _append_workbook_groundings(
            workbook_id,
            tool_name="add_sheet_from_shell",
            from_skill_ids=[shell_id],
            target_node=sheet_name,
            adaptation_notes=f"sheet shell kwargs={kwargs_json[:240]}",
            extra={"sheet_name": sheet_name},
        )
    except Exception as e:
        return _err(f"add_sheet_from_shell failed: {e}")
    return _ok(out)


@mcp.tool()
def init_from_archetype(workbook_id: str, archetype_id: str,
                        kwargs_json: str = "{}") -> str:
    """Initialize a multi-sheet workbook from a tier-5 archetype.

    Archetypes lay out the sheet structure (cover, data sheets, summary)
    and apply theme uniformly. Subsequent execute_xlsx_code or
    apply_component calls fill in task-specific content.
    """
    try:
        kwargs = json.loads(kwargs_json) if kwargs_json else {}
    except json.JSONDecodeError as e:
        return _err(f"kwargs_json invalid: {e}")
    try:
        out = engine.render_archetype(_skills(), workbook_id, archetype_id,
                                      kwargs)
        _append_workbook_groundings(
            workbook_id,
            tool_name="init_from_archetype",
            from_skill_ids=[archetype_id],
            target_node="workbook_archetype",
            adaptation_notes=f"archetype kwargs={kwargs_json[:240]}",
        )
    except Exception as e:
        return _err(f"init_from_archetype failed: {e}")
    return _ok(out)


# ---------------------------------------------------------------------------
# Skill library browsing
# ---------------------------------------------------------------------------

@mcp.tool()
def list_skills(tier: str = "", category: str = "",
                theme: str = "", limit: int = 50) -> str:
    """Browse the skill library. Filter by tier (component / sheet_shell /
    archetype / etc), category, or theme."""
    idx = _load_index()
    skills = idx.get("skills", [])
    out = []
    for s in skills:
        if tier and s.get("tier") != tier:
            continue
        if category and s.get("category") != category:
            continue
        if theme and s.get("theme") != theme:
            continue
        out.append({
            "skill_id": s.get("skill_id"),
            "skill_name": s.get("skill_name"),
            "tier": s.get("tier"),
            "category": s.get("category"),
            "theme": s.get("theme"),
            "applicability": s.get("applicability", "")[:200],
        })
        if len(out) >= limit:
            break
    return _ok({"total": len(out), "skills": out,
                "library_total": idx.get("total", len(skills))})


@mcp.tool()
def get_skill_info(skill_id: str) -> str:
    """Return full metadata for a skill (description, applicability,
    parameters, source video if any)."""
    idx = _load_index()
    for s in idx.get("skills", []):
        if s.get("skill_id") == skill_id:
            detail_path = s.get("detail_path")
            if detail_path:
                p = _skills() / detail_path
                if p.exists():
                    with p.open("r", encoding="utf-8") as f:
                        return _ok(json.load(f))
            return _ok(s)
    return _err(f"skill_id {skill_id!r} not found in index")


@mcp.tool()
def get_skill_code(skill_id: str) -> str:
    """Return the source/code for a skill.

    For seed skills (.py modules under tier dirs) returns the file source.
    For distilled skills (collected via cli.py collect, stored as JSON with
    code embedded in the analysis markdown) extracts and returns the
    fenced ```python code block``` from the analysis.
    """
    import re
    idx = _load_index()
    for s in idx.get("skills", []):
        if s.get("skill_id") != skill_id:
            continue

        # 1. Tier-dir .py file (seed skills)
        tier = s.get("tier")
        tier_dir_map = {
            "component": "components",
            "sheet_shell": "sheet_shells_seed",
            "archetype": "workbook_archetypes",
        }
        tier_dir = tier_dir_map.get(tier)
        if tier_dir:
            py = _skills() / tier_dir / f"{skill_id}.py"
            if py.exists():
                return py.read_text(encoding="utf-8")

        # 2. Distilled skill — extract code from detail JSON's analysis markdown
        detail_rel = s.get("detail_path")
        if detail_rel:
            detail_path = _skills() / detail_rel
            if detail_path.exists():
                detail = json.loads(detail_path.read_text(encoding="utf-8"))
                analysis = detail.get("analysis", "")
                # Try python first, then json (for tokens/snippets)
                for lang in ("python", "json"):
                    m = re.search(
                        rf"```{lang}\s*\n(.*?)```",
                        analysis, re.DOTALL,
                    )
                    if m:
                        return m.group(1).rstrip() + "\n"
                # Fallback: any fenced block
                m = re.search(r"```[a-zA-Z]*\s*\n(.*?)```",
                              analysis, re.DOTALL)
                if m:
                    return m.group(1).rstrip() + "\n"
                return _err(f"no fenced code block in analysis for {skill_id}")
        return _err(f"no source available for {skill_id} "
                    f"(tier={tier!r}, detail_path={detail_rel!r})")
    return _err(f"skill_id {skill_id!r} not found")


# ---------------------------------------------------------------------------
# Entry
# ---------------------------------------------------------------------------

def main() -> None:
    global _skills_dir, _demo_root
    parser = argparse.ArgumentParser()
    parser.add_argument("--skills-dir", type=str, default=None,
                        help="Path to skills_library/excel/ (override). When omitted, "
                             "the skills dir is resolved from domain.yaml's library_backend.")
    parser.add_argument("--demo-dir", type=str, default=None,
                        help="Path to demo/excel/ (default: <project>/demo/excel)")
    args = parser.parse_args()

    if args.skills_dir:
        _skills_dir = Path(args.skills_dir).resolve()
    else:
        # Honour the per-domain library_backend flag rather than hard-coding
        # the legacy path; flipping domain.yaml + restarting the server is
        # the documented cutover.
        from core import get_library_dir
        _skills_dir = get_library_dir("excel").resolve()
    if args.demo_dir:
        _demo_root = Path(args.demo_dir).resolve()
    else:
        _demo_root = (_PROJECT_ROOT / "demo" / "excel").resolve()
    _demo_root.mkdir(parents=True, exist_ok=True)

    # When the wiki backend is active, expose the universal discovery +
    # apply_skill surface alongside the legacy Excel-specific tools so an
    # agent can browse and dispatch through the wiki contract.
    from core import get_active_library_backend
    backend = get_active_library_backend("excel")
    global _startup_backend, _wiki_tools_registered
    _startup_backend = backend
    if backend == "wiki":
        try:
            from domains.excel.wiki_adapter import ExcelWikiAdapter
            from core.skill_wiki.mcp_tools import register_wiki_tools
            register_wiki_tools(mcp, ExcelWikiAdapter())
            _wiki_tools_registered = True
            log.info("excel-mcp: registered universal wiki discovery tools")
        except Exception as exc:  # noqa: BLE001
            log.error("excel-mcp: failed to register wiki tools: %s", exc)
            raise

    # Always install the legacy stale-registry guard so a backend flip on
    # a process that started in either mode is detected and surfaces a
    # structured error rather than silently serving stale data. The guard
    # exempts ``reload_registry`` so the documented remediation path
    # remains callable after a flip.
    from core.skill_wiki.legacy_stale import register_legacy_stale_check
    wrapped = register_legacy_stale_check(mcp, domain="excel", startup_backend=backend)
    log.info("excel-mcp: stale-registry guard installed on %d tools", wrapped)

    log.info("excel-mcp ready, skills_dir=%s demo_root=%s backend=%s",
             _skills_dir, _demo_root, backend)
    mcp.run(transport="stdio")


if __name__ == "__main__":
    main()
