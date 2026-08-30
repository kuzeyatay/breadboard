"""
sales_report_quarterly sheet shell
----------------------------------
Quarterly sales sheet: 3-month data + variance + chart + Excel table.

Tier: sheet_shell
Inputs:
    title (str)
    months (list[str]) — exactly 3 month names, e.g. ["Jan", "Feb", "Mar"]
    revenue (list[float]) — 3 monthly revenue numbers
    target  (list[float]) — 3 monthly target numbers
    theme (str)
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from _helpers import (  # noqa: E402
    Alignment, Font, apply_zebra, auto_width, body_font, fill, hex_to_argb,
    header_font, load_theme, thin_border, title_font, write_header_row,
)

import importlib.util as _ilu

_DOMAIN_DIR = Path(__file__).resolve().parents[1]


def _load(comp: str):
    spec = _ilu.spec_from_file_location(
        f"_excel_comp_{comp}", _DOMAIN_DIR / "components" / f"{comp}.py")
    mod = _ilu.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore
    return mod


def render_sheet(wb, sheet_name: str, *, title: str,
                 months: list[str], revenue: list[float], target: list[float],
                 theme: str = "corporate_blue") -> None:
    if len(months) != 3 or len(revenue) != 3 or len(target) != 3:
        raise ValueError("sales_report_quarterly expects 3 months/revenue/target values")

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif (len(wb.sheetnames) == 1 and wb.active.max_row <= 1
          and wb.active.max_column <= 1):
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    th = load_theme(theme)
    border = thin_border(th["border_color"])

    # --- Title row ---
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value=title)
    t.font = title_font(th, size=th.get("font_size_title", 16) + 2)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # --- Subtitle band ---
    ws.merge_cells("A2:F2")
    s = ws.cell(row=2, column=1)
    s.fill = fill(th["accent"])
    ws.row_dimensions[2].height = 4

    # --- Headers row 4 ---
    headers = ["Month", "Revenue", "Target", "Variance", "Var %", "Status"]
    write_header_row(ws, row=4, headers=headers, theme=th)

    # --- Data rows 5-7 ---
    body_f = body_font(th)
    for i, (m, r, t_) in enumerate(zip(months, revenue, target)):
        row = 5 + i
        ws.cell(row=row, column=1, value=m).font = body_f
        ws.cell(row=row, column=2, value=r).number_format = "#,##0"
        ws.cell(row=row, column=3, value=t_).number_format = "#,##0"
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "+#,##0;-#,##0"
        ws.cell(row=row, column=5,
                value=f"=IFERROR((B{row}-C{row})/C{row},0)").number_format = "+0.0%;-0.0%"
        ws.cell(row=row, column=6,
                value=f"=IF(B{row}>=C{row},\"On Target\",\"Below\")")
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=c).font = body_f

    apply_zebra(ws, 5, 7, 1, 6, th)

    # --- Summary row 8 ---
    sumrow_mod = _load("summary_row")
    sumrow_mod.render(
        ws, anchor="A5",
        data_range="A4:F7",
        label="TOTAL", op="SUM",
        number_format="#,##0", theme=theme,
    )
    # Override col E summary to be average %
    ws.cell(row=8, column=5,
            value="=IFERROR(AVERAGE(E5:E7),0)").number_format = "+0.0%;-0.0%"
    ws.cell(row=8, column=6, value="").value = ""  # leave status blank

    # --- Chart in cols H+ ---
    chart_mod = _load("compare_chart")
    chart_mod.render(ws, anchor="H4", data_range="A4:C7",
                     title=f"{title} — Revenue vs Target",
                     template="bar_compare", theme=theme,
                     sheet_name=sheet_name)

    # --- Excel Table over A4:F7 ---
    from openpyxl.worksheet.table import Table, TableStyleInfo
    table_id = f"tbl_{sheet_name.replace(' ', '_')}"
    tbl = Table(displayName=table_id, ref="A4:F7")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    try:
        ws.add_table(tbl)
    except Exception:
        pass

    auto_width(ws, 1, 6, padding=3, hard_max=22)
