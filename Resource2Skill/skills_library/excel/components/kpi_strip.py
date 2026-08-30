"""
kpi_strip component
-------------------
Renders a horizontal row of KPI cards. Each card has a label, a big value,
and an optional delta (vs prior or vs target).

Tier: component
Inputs:
    metrics (list[dict]): [{"label": "Revenue", "value": "$1.2M",
                            "delta": "+12.4%", "delta_kind": "good"|"bad"|"warn"|None},
                           ...]
    theme (str): palette name
    card_width (int): columns per card (default 3)
    card_height_rows (int): rows per card (default 4)
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from _helpers import (  # noqa: E402
    Alignment, Font, body_font, fill, hex_to_argb, load_theme, thin_border,
)


def render(ws, anchor: str, *, metrics: list[dict],
           theme: str = "corporate_blue", card_width: int = 3,
           card_height_rows: int = 4) -> None:
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    col_letter, row = coordinate_from_string(anchor)
    start_col = column_index_from_string(col_letter)

    th = load_theme(theme)
    label_font = Font(name=th.get("font_body", "Calibri"),
                      size=10, color=hex_to_argb(th["body_fg"]),
                      italic=False, bold=False)
    value_font = Font(name=th.get("font_title", "Calibri"),
                      size=20, bold=True,
                      color=hex_to_argb(th["title_fg"]))

    delta_color_map = {
        "good": th["good"],
        "bad":  th["bad"],
        "warn": th["warn"],
        None:   th["body_fg"],
    }

    card_bg = fill(th["zebra_bg"])
    border = thin_border(th["border_color"])

    for i, m in enumerate(metrics):
        c0 = start_col + i * (card_width + 1)
        c1 = c0 + card_width - 1
        # Card background spanning card rows
        for rr in range(row, row + card_height_rows):
            for cc in range(c0, c1 + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = card_bg
                cell.border = border
        ws.row_dimensions[row + 0].height = 18
        ws.row_dimensions[row + 1].height = 28
        ws.row_dimensions[row + 2].height = 18

        # Label
        ws.merge_cells(start_row=row, start_column=c0,
                       end_row=row, end_column=c1)
        lab = ws.cell(row=row, column=c0, value=m.get("label", ""))
        lab.font = label_font
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Value
        ws.merge_cells(start_row=row + 1, start_column=c0,
                       end_row=row + 1, end_column=c1)
        val = ws.cell(row=row + 1, column=c0, value=m.get("value", ""))
        val.font = value_font
        val.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Delta
        if m.get("delta"):
            kind = m.get("delta_kind")
            ws.merge_cells(start_row=row + 2, start_column=c0,
                           end_row=row + 2, end_column=c1)
            d = ws.cell(row=row + 2, column=c0, value=m["delta"])
            d.font = Font(name=th.get("font_body", "Calibri"),
                          size=11, bold=True,
                          color=hex_to_argb(delta_color_map.get(kind, th["body_fg"])))
            d.alignment = Alignment(horizontal="left", vertical="center", indent=1)
