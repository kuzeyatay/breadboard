"""
summary_row component
---------------------
Append a SUM/AVG/MIN/MAX summary row beneath an existing data range.

Tier: component
Inputs:
    data_range (str): e.g. "B2:E13" — last row of data
    label_col (int): column index for the row label (default same col as start)
    label (str): "TOTAL" / "AVERAGE" / "MEAN" / etc.
    op (str): "SUM" | "AVERAGE" | "MIN" | "MAX"
    number_format (str): optional format for summary cells
    theme (str): palette name
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from _helpers import (  # noqa: E402
    Alignment, Font, bottom_border, fill, hex_to_argb, load_theme,
)


def render(ws, anchor: str, *, data_range: str,
           label: str = "TOTAL", op: str = "SUM",
           number_format: str = "#,##0",
           theme: str = "corporate_blue") -> None:
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    op_u = op.upper()
    if op_u not in {"SUM", "AVERAGE", "MIN", "MAX"}:
        raise ValueError(f"unsupported op {op!r}")

    min_col, min_row, max_col, max_row = range_boundaries(data_range)
    summary_row = max_row + 1

    th = load_theme(theme)

    label_cell = ws.cell(row=summary_row, column=min_col, value=label)
    label_cell.font = Font(name=th.get("font_body", "Calibri"),
                           size=th.get("font_size_body", 11), bold=True,
                           color=hex_to_argb(th["body_fg"]))
    label_cell.fill = fill(th["zebra_bg"])
    label_cell.alignment = Alignment(horizontal="left", indent=1)
    label_cell.border = bottom_border(th["accent"], weight="medium")

    for c in range(min_col + 1, max_col + 1):
        col_letter = get_column_letter(c)
        formula = f"={op_u}({col_letter}{min_row + 1}:{col_letter}{max_row})"
        cell = ws.cell(row=summary_row, column=c, value=formula)
        cell.font = Font(name=th.get("font_body", "Calibri"),
                         size=th.get("font_size_body", 11), bold=True,
                         color=hex_to_argb(th["title_fg"]))
        cell.fill = fill(th["zebra_bg"])
        cell.number_format = number_format
        cell.border = bottom_border(th["accent"], weight="medium")
        cell.alignment = Alignment(horizontal="right")
