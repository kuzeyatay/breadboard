"""
_helpers.py — shared utilities for excel seed skills (components, shells,
archetypes). Loaded by importing siblings via relative path tricks; for
simplicity each skill imports this module directly when needed.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THEMES_CACHE: dict[str, dict] = {}


def _skills_root() -> Path:
    # This file lives at skills_library/excel/_helpers.py
    return Path(__file__).resolve().parent


def load_theme(name: str) -> dict:
    if name in _THEMES_CACHE:
        return _THEMES_CACHE[name]
    p = _skills_root() / "themes" / f"{name}.json"
    if not p.exists():
        # Fallback to corporate_blue if requested theme missing
        p = _skills_root() / "themes" / "corporate_blue.json"
    with p.open("r", encoding="utf-8") as f:
        data = json.load(f)
    _THEMES_CACHE[name] = data
    return data


def hex_to_argb(hex_color: str) -> str:
    h = hex_color.lstrip("#").upper()
    if len(h) == 6:
        return "FF" + h
    if len(h) == 8:
        return h
    return "FF000000"


def fill(theme_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_to_argb(theme_color))


def thin_border(color: str = "BFBFBF") -> Border:
    s = Side(border_style="thin", color=hex_to_argb(color))
    return Border(left=s, right=s, top=s, bottom=s)


def bottom_border(color: str = "1F4E78", weight: str = "medium") -> Border:
    s = Side(border_style=weight, color=hex_to_argb(color))
    return Border(bottom=s)


def header_font(theme: dict, size: int | None = None) -> Font:
    return Font(
        name=theme.get("font_body", "Calibri"),
        size=size or theme.get("font_size_subheader", 12),
        bold=True,
        color=hex_to_argb(theme["header_fg"]),
    )


def title_font(theme: dict, size: int | None = None) -> Font:
    return Font(
        name=theme.get("font_title", "Calibri"),
        size=size or theme.get("font_size_title", 16),
        bold=True,
        color=hex_to_argb(theme["title_fg"]),
    )


def body_font(theme: dict) -> Font:
    return Font(
        name=theme.get("font_body", "Calibri"),
        size=theme.get("font_size_body", 11),
        color=hex_to_argb(theme["body_fg"]),
    )


def auto_width(ws, min_col: int = 1, max_col: int | None = None,
               padding: int = 2, hard_max: int = 38) -> None:
    last_col = max_col or ws.max_column
    for col_idx in range(min_col, last_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[letter].width = min(max_len + padding, hard_max)


def apply_zebra(ws, start_row: int, end_row: int,
                start_col: int, end_col: int, theme: dict) -> None:
    z = fill(theme["zebra_bg"])
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).fill = z


def write_header_row(ws, row: int, headers: list[str], theme: dict,
                     start_col: int = 1) -> None:
    bg = fill(theme["header_bg"])
    fnt = header_font(theme)
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = fnt
        c.fill = bg
        c.alignment = align
    ws.row_dimensions[row].height = 22
