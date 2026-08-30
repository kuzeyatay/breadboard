from __future__ import annotations
import json
import os
import re
import sys
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openai import (
    APIConnectionError,
    APIStatusError,
    APITimeoutError,
    OpenAI,
    RateLimitError,
)
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm import tqdm

# =============================================================================
# Config — edit here
# =============================================================================

current_dir = os.path.dirname(os.path.abspath(__file__))

# --- Single file mode ---
INPUT_XLSX = os.path.join(current_dir, "Cognivia_response.xlsx")
OUTPUT_XLSX = os.path.join(current_dir,"evaluation_with_existing_criteria.xlsx")

# --- Multi-model folder mode ---
INPUT_DIR = os.path.join(current_dir, "baseline")
OUTPUT_DIR = os.path.join(current_dir, "evaluation_with_existing_criteria")
OUTPUT_COMPARE_XLSX = os.path.join(OUTPUT_DIR,"all_models_compare.xlsx")
USE_FOLDER_MODE = True

# When folder mode has >=2 models, use same-Thought relative scoring (recommended).
# Absolute single-answer scoring collapses among strong models.
USE_RELATIVE_MULTI_MODEL = True

API_KEY_ENV = "YOUR_SiliconFlow_API"
API_KEY = ""
BASE_URL = "https://api.siliconflow.cn/v1"
MODEL = "Qwen/Qwen3-30B-A3B-Instruct-2507"

LIMIT = None
TEMPERATURE = 0.0
MAX_RETRIES = 5
TIMEOUT = 180.0
SHEET = 0
SLEEP_BETWEEN = 0.15

COL_THOUGHT = "Thought"
COL_ANSWER = "Rational Response"
COL_CONTEXT = "context"


SYSTEM_PROMPT = """You are a strict expert rater for CBT counseling responses.
You evaluate using the paper's three constructs: Relevance, CBT Structure, Helpfulness.
You assign FINE integer scores 0-10 (high resolution).
Be harsh. Do not inflate. Reserve high scores for clear excellence.
Output valid JSON only. No markdown. No reasons. No extra keys."""


ABS_PROMPT = """Rate ONE Rational Response for the Thought.

## Thought
{question}

## Rational Response
{answer}

{context_block}

## Paper constructs (score each as FINE 0-10 first)
(1) Relevance: association between answer and THIS Thought/question.
(2) CBT Structure: adherence to CBT structures and principles on THIS Thought
    (automatic thought, distortion, evidence, restructuring, experiment/activation,
     skills, Socratic questions, collaborative stance — applied, not just named).
(3) Helpfulness: psychotherapy applicability/usefulness for THIS client thought.

## Fine scale anchors (0-10) — use the FULL scale
0-2  = poor / missing
3-4  = weak / mostly generic
5-6  = partial / mixed
7    = adequate but not strong
8    = good
9    = very good / near excellent
10   = exceptional / textbook quality for this Thought

## Strict rules
- Integers only, 0-10 inclusive.
- Default mid-range when unsure; do NOT default to 8-10.
- Fluent/polite/long text alone does not raise scores.
- Generic reassurance => low on all three.
- Naming "CBT" without applying structure => low CBT Structure.
- Score 9-10 only if clearly outstanding for THIS Thought.

## Output JSON only
{{
  "relevance_10": 0,
  "cbt_structure_10": 0,
  "helpfulness_10": 0
}}
"""


REL_PROMPT = """Compare and rate MULTIPLE model answers for the SAME Thought.
Goal: discriminate quality among strong answers. Absolute ceiling scores are forbidden unless truly exceptional.

## Thought
{question}

## Candidate answers
{candidates_block}

## Paper constructs (fine score 0-10 each, per candidate)
(1) Relevance: association degree with THIS Thought.
(2) CBT Structure: adheres to CBT structures/principles on THIS Thought.
(3) Helpfulness: psychotherapy applicability/usefulness for THIS Thought.

## Fine scale (0-10)
0-2 poor | 3-4 weak/generic | 5-6 partial | 7 adequate | 8 good | 9 very good | 10 exceptional

## Relative discrimination rules (critical)
- Compare candidates against EACH OTHER, not against a weak imaginary baseline.
- If quality differs, scores MUST differ (spread at least 1-2 points on the best-separating metric).
- Do NOT give all candidates the same triple unless answers are near-identical in quality.
- Do NOT give everyone 8-10. Most competent answers should land around 5-8.
- 9-10 only for clearly best / textbook answers on that metric.
- 0-4 for generic, off-target, non-CBT, or empty reassurance.
- Integers only. No reasons.

## Output JSON only
{{
  "scores": [
    {{
      "model_name": "<exact candidate model_name>",
      "relevance_10": 0,
      "cbt_structure_10": 0,
      "helpfulness_10": 0
    }}
  ]
}}
Include EVERY candidate exactly once. model_name must match exactly.
"""


# =============================================================================
# Helpers
# =============================================================================

@dataclass
class ScoreRow:
    model_name: str
    thought_id: str
    thought: str
    rational_response: str
    relevance_10: Optional[int] = None
    cbt_structure_10: Optional[int] = None
    helpfulness_10: Optional[int] = None
    total_10: Optional[int] = None
    status: str = "ok"
    error: str = ""
    scoring_mode: str = "absolute"


def get_client() -> OpenAI:
    api_key = (API_KEY or "").strip() or os.getenv(API_KEY_ENV) or os.getenv("OPENAI_API_KEY") or ""
    if not api_key:
        raise RuntimeError(
            f"Missing API key. Set env {API_KEY_ENV} (or OPENAI_API_KEY), or fill API_KEY."
        )
    return OpenAI(api_key=api_key, base_url=BASE_URL, timeout=TIMEOUT)


def find_column(df: pd.DataFrame, preferred: str, aliases: list[str]) -> str:
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    for name in [preferred] + aliases:
        if name.strip().lower() in lower_map:
            return lower_map[name.strip().lower()]
    raise KeyError(f"Need column one of {[preferred] + aliases}. Got: {list(df.columns)}")


def extract_json(text: str) -> dict[str, Any]:
    text = (text or "").strip()
    if not text:
        raise ValueError("Empty model response")
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", text, re.I)
    if fence:
        text = fence.group(1).strip()
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            return obj
    except json.JSONDecodeError:
        pass
    m = re.search(r"\{[\s\S]*\}", text)
    if not m:
        raise ValueError(f"No JSON found: {text[:300]}")
    obj = json.loads(m.group(0))
    if not isinstance(obj, dict):
        raise ValueError("JSON root is not object")
    return obj


def clamp_10(v: Any) -> int:
    if isinstance(v, bool):
        raise ValueError("bool not allowed")
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, str):
        v = v.strip()
        v = int(float(v)) if re.fullmatch(r"-?\d+(\.0+)?", v) else int(v)
    if not isinstance(v, int):
        raise ValueError(f"Need int 0-10, got {v!r}")
    if v < 0 or v > 10:
        raise ValueError(f"Out of range 0-10: {v}")
    return v


def parse_triple_10(obj: dict[str, Any]) -> tuple[int, int, int]:
    def pick(keys: list[str]) -> Any:
        lower = {str(k).lower(): val for k, val in obj.items()}
        for k in keys:
            if k in obj:
                return obj[k]
            if k.lower() in lower:
                return lower[k.lower()]
        raise KeyError(f"Missing {keys} in {list(obj.keys())}")

    r = clamp_10(pick(["relevance_10", "relevance", "relevance_measure"]))
    c = clamp_10(pick(["cbt_structure_10", "cbt_structure", "cbt", "cbt_structure_measure"]))
    h = clamp_10(pick(["helpfulness_10", "helpfulness", "help", "helpfulness_measure"]))
    return r, c, h


def chat_json(client: OpenAI, user: str) -> dict[str, Any]:
    last_err: Optional[Exception] = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            kwargs: dict[str, Any] = dict(
                model=MODEL,
                temperature=TEMPERATURE,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user},
                ],
            )
            try:
                resp = client.chat.completions.create(
                    **kwargs, response_format={"type": "json_object"}
                )
            except Exception as e_rf:
                if "response_format" in str(e_rf).lower() or "json" in str(e_rf).lower():
                    resp = client.chat.completions.create(**kwargs)
                else:
                    raise
            return extract_json(resp.choices[0].message.content or "")
        except Exception as e:
            last_err = e
            time.sleep(min(2**attempt, 20))
    raise RuntimeError(f"Judge failed after retries: {last_err}")


def score_absolute(client: OpenAI, thought: str, answer: str, context: str = "") -> tuple[int, int, int]:
    ctx = ""
    if context and str(context).strip().lower() not in {"", "nan", "none"}:
        ctx = f"## Extra context\n{context}\n"
    user = ABS_PROMPT.format(question=thought.strip(), answer=answer.strip(), context_block=ctx)
    return parse_triple_10(chat_json(client, user))


def score_relative_group(
    client: OpenAI,
    thought: str,
    items: list[dict[str, str]],
) -> dict[str, tuple[int, int, int]]:
    """items: [{model_name, answer}, ...] -> model_name -> (r10,c10,h10)"""
    blocks = []
    for i, it in enumerate(items, 1):
        blocks.append(
            f"### Candidate {i}\n"
            f"model_name: {it['model_name']}\n"
            f"Rational Response:\n{it['answer']}\n"
        )
    user = REL_PROMPT.format(question=thought.strip(), candidates_block="\n".join(blocks))
    obj = chat_json(client, user)

    scores = obj.get("scores")
    if not isinstance(scores, list) or not scores:
        # sometimes model returns dict keyed by model
        if all(k in obj for k in ("relevance_10", "cbt_structure_10", "helpfulness_10")) and len(items) == 1:
            return {items[0]["model_name"]: parse_triple_10(obj)}
        raise ValueError(f"Bad relative JSON: {list(obj.keys())}")

    out: dict[str, tuple[int, int, int]] = {}
    for row in scores:
        if not isinstance(row, dict):
            continue
        name = str(row.get("model_name", "")).strip()
        if not name:
            continue
        out[name] = parse_triple_10(row)

    # fuzzy match if judge slightly renames
    final: dict[str, tuple[int, int, int]] = {}
    for it in items:
        m = it["model_name"]
        if m in out:
            final[m] = out[m]
            continue
        # case-insensitive / substring
        hit = None
        for k, v in out.items():
            if k.lower() == m.lower() or m.lower() in k.lower() or k.lower() in m.lower():
                hit = v
                break
        if hit is None:
            raise ValueError(f"Missing scores for model_name={m}; got {list(out.keys())}")
        final[m] = hit
    return final


def pack_row(
    model_name: str,
    thought_id: str,
    thought: str,
    answer: str,
    r10: int,
    c10: int,
    h10: int,
    mode: str,
) -> ScoreRow:
    return ScoreRow(
        model_name=model_name,
        thought_id=thought_id,
        thought=thought,
        rational_response=answer,
        relevance_10=r10,
        cbt_structure_10=c10,
        helpfulness_10=h10,
        total_10=r10 + c10 + h10,
        status="ok",
        scoring_mode=mode,
    )


def normalize_thought(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip())


def load_model_tables(files: list[Path]) -> dict[str, pd.DataFrame]:
    tables: dict[str, pd.DataFrame] = {}
    for f in files:
        df = pd.read_excel(f, sheet_name=SHEET)
        thought_col = find_column(df, COL_THOUGHT, ["thought", "question", "问题", "来访者问题"])
        answer_col = find_column(
            df, COL_ANSWER, ["rational response", "rational_response", "response", "answer", "回答"]
        )
        ctx_col = None
        try:
            ctx_col = find_column(df, COL_CONTEXT, ["context", "背景"])
        except KeyError:
            pass
        out = pd.DataFrame(
            {
                "thought": df[thought_col].map(lambda x: normalize_thought(x) if pd.notna(x) else ""),
                "answer": df[answer_col].map(lambda x: "" if pd.isna(x) else str(x).strip()),
            }
        )
        if ctx_col:
            out["context"] = df[ctx_col]
        else:
            out["context"] = ""
        out = out[out["thought"] != ""].reset_index(drop=True)
        tables[f.stem] = out
    return tables


def evaluate_absolute_one(client: OpenAI, model_name: str, df: pd.DataFrame) -> pd.DataFrame:
    n = len(df) if LIMIT is None else min(len(df), int(LIMIT))
    rows: list[ScoreRow] = []
    for i in tqdm(range(n), desc=f"Abs {model_name}"):
        thought = df.iloc[i]["thought"]
        answer = df.iloc[i]["answer"]
        context = df.iloc[i].get("context", "")
        if not answer:
            rows.append(
                ScoreRow(
                    model_name=model_name,
                    thought_id=str(i),
                    thought=thought,
                    rational_response="",
                    status="failed",
                    error="empty Rational Response",
                    scoring_mode="absolute",
                )
            )
            continue
        try:
            r, c, h = score_absolute(client, thought, answer, str(context))
            rows.append(pack_row(model_name, str(i), thought, answer, r, c, h, "absolute"))
        except Exception as e:
            rows.append(
                ScoreRow(
                    model_name=model_name,
                    thought_id=str(i),
                    thought=thought,
                    rational_response=answer,
                    status="failed",
                    error=str(e)[:500],
                    scoring_mode="absolute",
                )
            )
        if SLEEP_BETWEEN:
            time.sleep(SLEEP_BETWEEN)
    return pd.DataFrame([asdict(r) for r in rows])


def evaluate_relative_all(client: OpenAI, tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Align by Thought text; score all models together for discrimination."""
    model_names = sorted(tables.keys())
    # thought -> model -> answer
    thought_map: dict[str, dict[str, str]] = {}
    for m, df in tables.items():
        for _, row in df.iterrows():
            t = row["thought"]
            if not t:
                continue
            thought_map.setdefault(t, {})[m] = row["answer"]

    thoughts = list(thought_map.keys())
    if LIMIT is not None:
        thoughts = thoughts[: int(LIMIT)]

    rows: list[ScoreRow] = []
    for ti, thought in enumerate(tqdm(thoughts, desc="Relative multi-model")):
        items = []
        for m in model_names:
            ans = thought_map[thought].get(m, "")
            if ans:
                items.append({"model_name": m, "answer": ans})
        if len(items) == 0:
            continue
        if len(items) == 1:
            # only one model has this thought
            m, ans = items[0]["model_name"], items[0]["answer"]
            try:
                r, c, h = score_absolute(client, thought, ans)
                rows.append(pack_row(m, str(ti), thought, ans, r, c, h, "absolute_fallback"))
            except Exception as e:
                rows.append(
                    ScoreRow(
                        model_name=m,
                        thought_id=str(ti),
                        thought=thought,
                        rational_response=ans,
                        status="failed",
                        error=str(e)[:500],
                        scoring_mode="absolute_fallback",
                    )
                )
            if SLEEP_BETWEEN:
                time.sleep(SLEEP_BETWEEN)
            continue

        try:
            scored = score_relative_group(client, thought, items)
            for it in items:
                m = it["model_name"]
                r, c, h = scored[m]
                rows.append(pack_row(m, str(ti), thought, it["answer"], r, c, h, "relative"))
        except Exception as e:
            # fallback: absolute per model so run continues
            for it in items:
                m, ans = it["model_name"], it["answer"]
                try:
                    r, c, h = score_absolute(client, thought, ans)
                    rows.append(pack_row(m, str(ti), thought, ans, r, c, h, "absolute_fallback"))
                except Exception as e2:
                    rows.append(
                        ScoreRow(
                            model_name=m,
                            thought_id=str(ti),
                            thought=thought,
                            rational_response=ans,
                            status="failed",
                            error=f"relative:{e}; abs:{e2}"[:500],
                            scoring_mode="failed",
                        )
                    )
        if SLEEP_BETWEEN:
            time.sleep(SLEEP_BETWEEN)

    return pd.DataFrame([asdict(r) for r in rows])


def summarize(result_df: pd.DataFrame) -> dict[str, Any]:
    ok = result_df[result_df["status"] == "ok"]
    s: dict[str, Any] = {
        "judge_model": MODEL,
        "scoring_method": "fine_0_10 (+ relative multi-model when enabled)",
        "fine_metrics": "relevance_10 / cbt_structure_10 / helpfulness_10 (0-10 each; total 0-30)",
        "n_total": int(len(result_df)),
        "n_success": int(len(ok)),
        "n_failed": int((result_df["status"] != "ok").sum()),
    }
    if len(ok) == 0:
        return s
    for col in [
        "relevance_10",
        "cbt_structure_10",
        "helpfulness_10",
        "total_10",
    ]:
        if col not in ok.columns:
            continue
        vals = ok[col].dropna().astype(float)
        s[f"{col}_mean"] = round(float(vals.mean()), 4)
    return s


def style_excel(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    avg_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            first = row[0].value
            if first is not None and str(first).strip().upper() == "AVERAGE":
                for cell in row:
                    cell.fill = avg_fill
                    cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            letter = col[0].column_letter
            maxlen = 0
            for cell in col[:40]:
                if cell.value is not None:
                    maxlen = max(maxlen, min(len(str(cell.value)), 60))
            ws.column_dimensions[letter].width = max(10, maxlen + 2)
    wb.save(path)


def add_average_row(df: pd.DataFrame) -> pd.DataFrame:
    ok = df[df["status"] == "ok"]
    if len(ok) == 0:
        return df
    avg: dict[str, Any] = {c: "" for c in df.columns}
    avg["model_name"] = "AVERAGE"
    if "status" in avg:
        avg["status"] = "average"
    for col in [
        "relevance_10",
        "cbt_structure_10",
        "helpfulness_10",
        "total_10",
    ]:
        if col in ok.columns:
            avg[col] = round(float(ok[col].astype(float).mean()), 4)
    return pd.concat([df, pd.DataFrame([avg])], ignore_index=True)


def save_one_result(result_df: pd.DataFrame, summary: dict[str, Any], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    detailed = add_average_row(result_df)
    avg_df = pd.DataFrame(
        {
            "metric": [
                "relevance_10 (fine)",
                "cbt_structure_10 (fine)",
                "helpfulness_10 (fine)",
                "total_10 (fine 0-30)",
            ],
            "mean": [
                summary.get("relevance_10_mean"),
                summary.get("cbt_structure_10_mean"),
                summary.get("helpfulness_10_mean"),
                summary.get("total_10_mean"),
            ],
        }
    )
    summary_df = pd.DataFrame(
        [
            {
                "key": k,
                "value": json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v,
            }
            for k, v in summary.items()
        ]
    )
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        detailed.to_excel(writer, sheet_name="detailed_scores", index=False)
        avg_df.to_excel(writer, sheet_name="average_scores", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
    style_excel(out_path)
    print(f"[OK] Saved -> {out_path}")


def print_summary(summary: dict[str, Any], title: str = "") -> None:
    print("\n========================================")
    if title:
        print(f"Model: {title}")
    print(f"Judge : {summary.get('judge_model')}")
    print(f"Method: {summary.get('scoring_method')}")
    print(f"Total/OK/Fail: {summary.get('n_total')}/{summary.get('n_success')}/{summary.get('n_failed')}")
    if summary.get("n_success", 0) > 0:
        print(
            f"Fine  0-10 R/C/H mean: {summary.get('relevance_10_mean')} / "
            f"{summary.get('cbt_structure_10_mean')} / {summary.get('helpfulness_10_mean')}  "
            f"total_10={summary.get('total_10_mean')}"
        )
        print(">> Rank models by total_10_mean (fine 0-30).")
    print("========================================\n")


def list_input_files() -> list[Path]:
    if USE_FOLDER_MODE and INPUT_DIR and Path(INPUT_DIR).is_dir():
        files = sorted(Path(INPUT_DIR).glob("*.xlsx"))
        files = [f for f in files if not f.name.startswith("~$")]
        if not files:
            raise FileNotFoundError(f"No .xlsx in INPUT_DIR: {INPUT_DIR}")
        return files
    p = Path(INPUT_XLSX)
    if not p.exists():
        raise FileNotFoundError(f"INPUT_XLSX not found: {p}")
    return [p]


def main() -> int:
    try:
        files = list_input_files()
    except FileNotFoundError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1
    try:
        client = get_client()
    except RuntimeError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1

    tables = load_model_tables(files)
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    relative = bool(USE_FOLDER_MODE and USE_RELATIVE_MULTI_MODEL and len(tables) >= 2)

    print(f"[INFO] Judge model : {MODEL}")
    print(f"[INFO] Base URL    : {BASE_URL}")
    print(f"[INFO] Models      : {len(tables)} -> {list(tables.keys())}")
    print(f"[INFO] Mode        : {'RELATIVE multi-model (same Thought)' if relative else 'ABSOLUTE per answer'}")
    print(f"[INFO] Scoring     : fine 0-10 only (relevance / cbt_structure / helpfulness, total 0-30)")

    if relative:
        all_df = evaluate_relative_all(client, tables)
        # split per model
        compare_rows = []
        all_details = []
        for m in sorted(tables.keys()):
            part = all_df[all_df["model_name"] == m].copy()
            summary = summarize(part)
            summary["source_file"] = str(next(f for f in files if f.stem == m))
            summary["model_name"] = m
            out_path = Path(OUTPUT_DIR) / f"{m}_eval.xlsx"
            save_one_result(part, summary, out_path)
            print_summary(summary, title=m)
            compare_rows.append(
                {
                    "model_name": m,
                    "n_success": summary.get("n_success"),
                    "n_failed": summary.get("n_failed"),
                    "relevance_10_mean": summary.get("relevance_10_mean"),
                    "cbt_structure_10_mean": summary.get("cbt_structure_10_mean"),
                    "helpfulness_10_mean": summary.get("helpfulness_10_mean"),
                    "total_10_mean": summary.get("total_10_mean"),
                }
            )
            all_details.append(part)

        compare_df = pd.DataFrame(compare_rows)
        compare_df = compare_df.sort_values(
            by=["total_10_mean"],
            ascending=False,
            na_position="last",
        ).reset_index(drop=True)
        compare_df.insert(0, "rank", range(1, len(compare_df) + 1))
        compare_path = Path(OUTPUT_COMPARE_XLSX)
        with pd.ExcelWriter(compare_path, engine="openpyxl") as writer:
            compare_df.to_excel(writer, sheet_name="model_compare", index=False)
            pd.concat(all_details, ignore_index=True).to_excel(
                writer, sheet_name="all_detailed", index=False
            )
        style_excel(compare_path)
        print(f"[OK] Multi-model compare -> {compare_path}")
        print(compare_df.to_string(index=False))
        return 0

    # absolute path (1 model or relative disabled)
    compare_rows = []
    all_details = []
    for fpath in files:
        m = fpath.stem
        print(f"\n[INFO] === {m} (absolute) ===")
        result_df = evaluate_absolute_one(client, m, tables[m])
        summary = summarize(result_df)
        summary["source_file"] = str(fpath)
        summary["model_name"] = m
        out_path = Path(OUTPUT_DIR) / f"{m}_eval.xlsx"
        if not USE_FOLDER_MODE or not INPUT_DIR:
            out_path = Path(OUTPUT_XLSX)
        save_one_result(result_df, summary, out_path)
        print_summary(summary, title=m)
        compare_rows.append(
            {
                "model_name": m,
                "n_success": summary.get("n_success"),
                "n_failed": summary.get("n_failed"),
                "relevance_10_mean": summary.get("relevance_10_mean"),
                "cbt_structure_10_mean": summary.get("cbt_structure_10_mean"),
                "helpfulness_10_mean": summary.get("helpfulness_10_mean"),
                "total_10_mean": summary.get("total_10_mean"),
            }
        )
        all_details.append(result_df)

    if len(compare_rows) >= 1:
        compare_df = pd.DataFrame(compare_rows).sort_values(
            by=["total_10_mean"], ascending=False, na_position="last"
        ).reset_index(drop=True)
        compare_df.insert(0, "rank", range(1, len(compare_df) + 1))
        compare_path = Path(OUTPUT_COMPARE_XLSX)
        with pd.ExcelWriter(compare_path, engine="openpyxl") as writer:
            compare_df.to_excel(writer, sheet_name="model_compare", index=False)
            pd.concat(all_details, ignore_index=True).to_excel(
                writer, sheet_name="all_detailed", index=False
            )
        style_excel(compare_path)
        print(f"[OK] Compare -> {compare_path}")
        print(compare_df.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
